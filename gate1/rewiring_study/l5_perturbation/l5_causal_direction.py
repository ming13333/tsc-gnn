"""
L5 — Public TF-perturbation directional causal support.

Design (non-circular, program-level, not edge-level):
  For each core TF recovered by the TSC-GNN rewiring (Sox10, Cebpb; Sox8/Gata2 as
  specificity controls), take its TARGET PROGRAM from the SAME mouse DoRothEA GRN
  used by the framework (independent of any perturbation data). Then test, in an
  INDEPENDENT public loss-of-function RNA-seq dataset, whether that target program
  is enriched among genes that go DOWN when the TF is knocked out/down.

  - Direction-consistent down-enrichment of a TF's own targets  => directional
    causal support for the TF -> program edge (module level, NOT individual edges).
  - Specificity: other TFs' target programs must be NULL in the same dataset.

Datasets:
  GSE269122 : oligodendrocyte-specific Sox10 conditional KO (corpus callosum, 4 ctrl / 4 Sox10-KO)
              supp = ctrl_vs_Sox10ko.xlsx (Ensembl ID, log2FC, p, padj)
  GSE273163 : Cebpb heterozygous KO Kupffer cells (liver, 3 WT / 3 Hete)
              supp = RAW counts (gene symbol, Hete_rep1-3, WT_rep1-3)
"""
import os, gzip, csv, math, random
import numpy as np
import openpyxl
import mygene

random.seed(20260709)
np.random.seed(20260709)

ROOT = r"C:\D 盘\科研\虚拟敲除\gate1"
L5 = ROOT + r"\rewiring_study\l5_perturbation"
DOROTHEA = ROOT + r"\data\dorothea\mouse_dorothea_regulon.tsv"

FOCUS_TFS = ["Sox10", "Sox8", "Cebpb", "Gata2"]   # Sox10/Sox8 from oligo study; Cebpb/Gata2 inflammation

# ---------------------------------------------------------------- load DoRothEA
def load_dorothea(path):
    tf_targets = {}
    all_tfs = set()
    with open(path) as f:
        r = csv.DictReader(f, delimiter="\t")
        for row in r:
            src, tgt = row["source"].strip(), row["target"].strip()
            if not src or not tgt:
                continue
            tf_targets.setdefault(src, set()).add(tgt)
            all_tfs.add(src)
    return tf_targets, all_tfs

tf_targets, all_tfs = load_dorothea(DOROTHEA)
print(f"[DoRothEA] {len(all_tfs)} TFs; focus sizes: " +
      ", ".join(f"{t}={len(tf_targets.get(t,[]))}" for t in FOCUS_TFS))

# ---------------------------------------------------------------- ensembl->symbol
_ens2sym = {}
def map_ensembl(ids):
    mg = mygene.MyGeneInfo()
    missing = [i for i in ids if i not in _ens2sym]
    if missing:
        for s in range(0, len(missing), 2000):
            chunk = missing[s:s+2000]
            try:
                res = mg.querymany(chunk, scopes="ensembl.gene", fields="symbol",
                                   species="mouse", as_dataframe=False, verbose=False)
                for x in res:
                    q = x.get("query")
                    sym = x.get("symbol")
                    if q is not None and sym:
                        _ens2sym[q] = sym if isinstance(sym, str) else sym[0] if isinstance(sym, list) else None
            except Exception as e:
                print("  mygene chunk err:", e)
    return {i: _ens2sym.get(i) for i in ids}

# ---------------------------------------------------------------- GSE269122 (Sox10 KO)
def load_gse269122():
    path = os.path.join(L5, "GSE269122_ctrl_vs_Sox10ko.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    idx = {h: i for i, h in enumerate(hdr)}
    ens, logfc, padj = [], [], []
    for row in rows:
        e = row[idx["GeneID"]]
        if e is None:
            continue
        lf = row[idx["log2(FC)"]]
        pa = row[idx["padj"]]
        if lf is None or (isinstance(lf, str) and lf == "NA"):
            continue
        ens.append(str(e))
        logfc.append(float(lf))
        padj.append(None if (pa is None or (isinstance(pa, str) and pa == "NA")) else float(pa))
    sym = map_ensembl(ens)
    gene_fc = {}
    for e, lf in zip(ens, logfc):
        s = sym.get(e)
        if s:
            gene_fc[s] = lf   # last wins (ENSMUSG unique -> fine)
    print(f"[GSE269122] {len(gene_fc)} genes mapped with log2FC")
    return gene_fc

# ---------------------------------------------------------------- GSE273163 (Cebpb KO)
def load_gse273163():
    path = os.path.join(L5, "GSE273163_CEBPB_KC_liver_RAW_counts.txt.gz")
    gene_fc = {}
    with gzip.open(path, "rt") as f:
        r = csv.reader(f, delimiter="\t")
        hdr = next(r)
        # cols: gene_id, Hete_rep1..3, WT_rep1..3
        hete = [i for i, h in enumerate(hdr) if h.startswith("Hete")]
        wt = [i for i, h in enumerate(hdr) if h.startswith("WT")]
        for row in r:
            g = row[0].strip()
            if not g or g.upper() in ("GENE_ID", "NA"):
                continue
            hv = [max(float(row[i]), 0) for i in hete]
            wv = [max(float(row[i]), 0) for i in wt]
            mh, mw = np.mean(hv) + 1, np.mean(wv) + 1
            if mh <= 0 or mw <= 0:
                continue
            gene_fc[g] = math.log2(mh / mw)
    print(f"[GSE273163] {len(gene_fc)} genes with log2FC")
    return gene_fc

# ---------------------------------------------------------------- stats
def enrichment(gene_fc, target_syms, n_perm=2000, down_thr=None):
    universe = list(gene_fc.keys())
    fc_arr = np.array([gene_fc[g] for g in universe])
    present = set(g for g in target_syms if g in gene_fc)
    if len(present) < 5:
        return None
    obs = np.mean([gene_fc[g] for g in present])
    # down-enrichment: activator KO -> targets more negative
    if down_thr is None:
        downs = set(g for g in universe if gene_fc[g] < 0)
    else:
        downs = set(g for g in universe if gene_fc[g] < down_thr)
    a = len(present & downs); b = len(present) - a
    c = len(downs) - a; d = len(universe) - len(downs) - b
    # Fisher one-sided (targets enriched in down)
    try:
        from scipy.stats import fisher_exact
        _, p_fish = fisher_exact([[a, b], [c, d]], alternative="greater")
    except Exception:
        p_fish = float("nan")
    # permutation on continuous mean log2FC (down direction)
    n = len(present)
    perm = np.empty(n_perm)
    univ_idx = np.arange(len(universe))
    for i in range(n_perm):
        samp = np.random.choice(univ_idx, size=n, replace=False)
        perm[i] = fc_arr[samp].mean()
    emp_down = (np.sum(perm <= obs) + 1) / (n_perm + 1)   # P(rand <= obs)
    emp_two = (np.sum(np.abs(perm) >= abs(obs)) + 1) / (n_perm + 1)
    return {
        "n_targets_present": len(present),
        "n_targets_total": len(target_syms),
        "mean_log2FC_targets": round(float(obs), 4),
        "mean_log2FC_background": round(float(fc_arr.mean()), 4),
        "n_down_overlap": a,
        "OR_down": round(float((a*d)/(b*c+1e-9)), 3) if (b*c) > 0 else None,
        "fisher_p_down": round(float(p_fish), 3) if p_fish == p_fish else None,
        "perm_emp_p_down": round(float(emp_down), 4),
        "perm_emp_p_two": round(float(emp_two), 4),
    }

def rank_among_all_tfs(gene_fc, tf_targets, focus):
    """rank of focus TF's target-set mean log2FC among ALL TF target sets (1=most negative)."""
    universe = set(gene_fc.keys())
    means = {}
    for tf, tg in tf_targets.items():
        present = [g for g in tg if g in universe]
        if len(present) >= 5:
            means[tf] = np.mean([gene_fc[g] for g in present])
    obs = means.get(focus)
    if obs is None:
        return None
    ranked = sorted(means.items(), key=lambda kv: kv[1])  # ascending: most negative first
    rank = [t for t, _ in ranked].index(focus) + 1
    return {"rank": rank, "n_tested": len(ranked),
            "mean_log2FC": round(float(obs), 4)}

# ---------------------------------------------------------------- run
results = {}

gene_fc_269 = load_gse269122()
gene_fc_273 = load_gse273163()

print("\n===== GSE269122 : oligodendrocyte Sox10 conditional KO (ctrl vs Sox10-KO) =====")
res269 = {}
for tf in FOCUS_TFS + ["Sox2", "Olig2"]:
    tgt = tf_targets.get(tf, set())
    if not tgt:
        continue
    e = enrichment(gene_fc_269, tgt, n_perm=2000)
    if e is None:
        continue
    rk = rank_among_all_tfs(gene_fc_269, tf_targets, tf)
    entry = dict(e); entry["rank"] = rk
    res269[tf] = entry
    if e:
        print(f"  {tf:6s} n={e['n_targets_present']:4d} meanFC={e['mean_log2FC_targets']:+.3f} "
              f"ORdown={e['OR_down']} fishP={e['fisher_p_down']} empP_down={e['perm_emp_p_down']} "
              f"rank={rk['rank']}/{rk['n_tested']}")
results["GSE269122_Sox10KO"] = res269

print("\n===== GSE273163 : Cebpb heterozygous KO Kupffer (WT vs Hete) =====")
res273 = {}
for tf in FOCUS_TFS:
    tgt = tf_targets.get(tf, set())
    if not tgt:
        continue
    e = enrichment(gene_fc_273, tgt, n_perm=2000)
    if e is None:
        continue
    rk = rank_among_all_tfs(gene_fc_273, tf_targets, tf)
    entry = dict(e); entry["rank"] = rk
    res273[tf] = entry
    if e:
        print(f"  {tf:6s} n={e['n_targets_present']:4d} meanFC={e['mean_log2FC_targets']:+.3f} "
              f"ORdown={e['OR_down']} fishP={e['fisher_p_down']} empP_down={e['perm_emp_p_down']} "
              f"rank={rk['rank']}/{rk['n_tested']}")
results["GSE273163_CebpbKO"] = res273

# ---------------------------------------------------------------- save
out = os.path.join(L5, "l5_causal_direction.json")
import json
with open(out, "w") as f:
    json.dump(results, f, indent=2)
print(f"\n[saved] {out}")
